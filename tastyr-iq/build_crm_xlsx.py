#!/usr/bin/env python3
"""Tastyr IQ — Professional CRM Workbook (XLSX) for Co-Founder Handoff"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import csv, os

BASE = os.path.dirname(os.path.abspath(__file__))

# ── Brand Colors ──
NAVY_FILL = PatternFill(start_color='0A1A2F', end_color='0A1A2F', fill_type='solid')
GOLD_FILL = PatternFill(start_color='D4A534', end_color='D4A534', fill_type='solid')
LIGHT_BG = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
RED_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
GREEN_HEADER = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
GOLD_HEADER = PatternFill(start_color='D4A534', end_color='D4A534', fill_type='solid')
GRAY_HEADER = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
BLUE_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Calibri', size=9, color='1A1A1A')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='0A1A2F')
SUBTITLE_FONT = Font(name='Calibri', size=11, bold=True, color='3A7CD5')
STAT_FONT = Font(name='Calibri', size=18, bold=True, color='0A1A2F')
STAT_LABEL_FONT = Font(name='Calibri', size=9, color='888888')
LINK_FONT = Font(name='Calibri', size=9, color='3A7CD5', underline='single')

THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# Column definitions matching CRM
CRM_COLUMNS = [
    'Segment', 'Name', 'Title/Role', 'Organization', 'City/State',
    'Email', 'LinkedIn', 'Other Social', 'Phone',
    'Investment Focus / Expertise', 'Check Size', 'Notable Work', 'Food/Tastyr Connection',
    'Why Contact This Person', 'Psychological Angle', 'Emotional Trigger',
    'Connection Degree', 'Warm Intro Path', 'Rank', 'Chain Strategy',
    'Approach Strategy', 'Custom Message Draft', 'Status', 'Notes'
]

# Optimal column widths
COL_WIDTHS = {
    'Segment': 14, 'Name': 22, 'Title/Role': 25, 'Organization': 22, 'City/State': 18,
    'Email': 25, 'LinkedIn': 35, 'Other Social': 20, 'Phone': 15,
    'Investment Focus / Expertise': 35, 'Check Size': 16, 'Notable Work': 40, 'Food/Tastyr Connection': 35,
    'Why Contact This Person': 45, 'Psychological Angle': 35, 'Emotional Trigger': 30,
    'Connection Degree': 14, 'Warm Intro Path': 35, 'Rank': 8, 'Chain Strategy': 25,
    'Approach Strategy': 35, 'Custom Message Draft': 50, 'Status': 14, 'Notes': 30
}

def load_crm():
    with open(os.path.join(BASE, "crm", "master-crm.csv"), "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def add_headers(ws, columns, fill=NAVY_FILL):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col_name, 15)

def add_contact_rows(ws, contacts, start_row=2):
    for row_idx, contact in enumerate(contacts, start_row):
        rank = int(contact.get('Rank', 0))

        # Row background by tier
        if rank >= 70:
            row_fill = GREEN_FILL
        elif rank >= 40:
            row_fill = YELLOW_FILL
        else:
            row_fill = RED_FILL

        for col_idx, col_name in enumerate(CRM_COLUMNS, 1):
            value = contact.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            cell.fill = row_fill

            # Special formatting for Rank column
            if col_name == 'Rank':
                cell.alignment = CENTER_ALIGN
                cell.font = Font(name='Calibri', size=10, bold=True, color='0A1A2F')

            # LinkedIn as hyperlink
            if col_name == 'LinkedIn' and value and value.startswith('http'):
                cell.font = LINK_FONT

        # Set row height
        ws.row_dimensions[row_idx].height = 45

def build_dashboard(ws, contacts):
    """Build a professional dashboard sheet."""
    total = len(contacts)
    seg_counts = {}
    for c in contacts:
        seg_counts[c["Segment"]] = seg_counts.get(c["Segment"], 0) + 1
    tier1 = sum(1 for c in contacts if int(c["Rank"]) >= 70)
    tier2 = sum(1 for c in contacts if 40 <= int(c["Rank"]) < 70)
    tier3 = sum(1 for c in contacts if int(c["Rank"]) < 40)

    # Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = 'TASTYR IQ — CRM DASHBOARD'
    cell.font = Font(name='Calibri', size=20, bold=True, color='FFFFFF')
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45

    # Subtitle
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = f'{total} Contacts | 12 Segments | Tri-State Area | March 2026'
    cell.font = Font(name='Calibri', size=11, color='D4A534')
    cell.fill = NAVY_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # Stat boxes row
    row = 4
    stats = [
        ('Total Contacts', str(total)),
        ('Tier 1 (70-100)', str(tier1)),
        ('Tier 2 (40-69)', str(tier2)),
        ('Tier 3 (1-39)', str(tier3)),
    ]
    for i, (label, value) in enumerate(stats):
        col = i * 2 + 1
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = STAT_FONT
        cell.alignment = CENTER_ALIGN
        fills = [BLUE_FILL, GREEN_FILL, YELLOW_FILL, RED_FILL]
        cell.fill = fills[i]

        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        cell = ws.cell(row=row+1, column=col, value=label)
        cell.font = STAT_LABEL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = fills[i]

    ws.row_dimensions[row].height = 35
    ws.row_dimensions[row+1].height = 20

    # Segment breakdown
    row = 7
    ws.cell(row=row, column=1, value='CONTACTS BY SEGMENT').font = SUBTITLE_FONT
    row = 8
    ws.cell(row=row, column=1, value='Segment').font = HEADER_FONT
    ws.cell(row=row, column=1).fill = NAVY_FILL
    ws.cell(row=row, column=2, value='Count').font = HEADER_FONT
    ws.cell(row=row, column=2).fill = NAVY_FILL
    ws.cell(row=row, column=3, value='% of Total').font = HEADER_FONT
    ws.cell(row=row, column=3).fill = NAVY_FILL

    for seg in sorted(seg_counts.keys(), key=lambda x: seg_counts[x], reverse=True):
        row += 1
        ws.cell(row=row, column=1, value=seg).font = BODY_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=seg_counts[seg]).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER
        pct = seg_counts[seg] / total
        ws.cell(row=row, column=3, value=pct).font = BODY_FONT
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER
        if row % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = LIGHT_BG

    # Total row
    row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True, color='0A1A2F')
    ws.cell(row=row, column=1).fill = GOLD_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=total).font = Font(name='Calibri', size=10, bold=True, color='0A1A2F')
    ws.cell(row=row, column=2).fill = GOLD_FILL
    ws.cell(row=row, column=2).alignment = CENTER_ALIGN
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value=1.0).font = Font(name='Calibri', size=10, bold=True, color='0A1A2F')
    ws.cell(row=row, column=3).fill = GOLD_FILL
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = THIN_BORDER

    # Chain strategies
    row += 2
    ws.cell(row=row, column=1, value='SOCIAL CHAIN STRATEGIES').font = SUBTITLE_FONT
    row += 1
    chains = [
        ('A: Restaurateur Gateway', 'Blogger > Restaurant > Investor'),
        ('B: University Pipeline', 'Faculty > Students > Alumni Angels'),
        ('C: FOMO Cascade', 'Influencers > Media > Market Pull'),
        ('D: NJ-to-NYC Bridge', 'Princeton > NJ Angels > NYC VCs'),
        ('E: Accelerator Fast-Track', 'Apply > Accept > Demo Day'),
        ('F: Food Industry Backdoor', 'POS/Distributor > Restaurants > Investors'),
        ('G: Media Momentum', 'Eater > Technical.ly > PhillyMag > Inquirer'),
        ('H: Connector Express', 'Super-Connectors > Multiple Intros'),
    ]
    ws.cell(row=row, column=1, value='Chain').font = HEADER_FONT
    ws.cell(row=row, column=1).fill = NAVY_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value='Flow').font = HEADER_FONT
    ws.cell(row=row, column=2).fill = NAVY_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    for chain_name, flow in chains:
        row += 1
        ws.cell(row=row, column=1, value=chain_name).font = BODY_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=flow).font = BODY_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        if row % 2 == 0:
            ws.cell(row=row, column=1).fill = LIGHT_BG
            ws.cell(row=row, column=2).fill = LIGHT_BG

    # Ranking formula
    row += 2
    ws.cell(row=row, column=1, value='RANKING FORMULA').font = SUBTITLE_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='RANK = (Investor Potential × 0.30) + (Network Reach × 0.25) + (FOMO Multiplier × 0.20) + (Accessibility × 0.15) + (Timing × 0.10)').font = BODY_FONT

    # Column widths for dashboard
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

def build_xlsx():
    contacts = load_crm()
    sorted_contacts = sorted(contacts, key=lambda r: int(r["Rank"]), reverse=True)
    tier1 = [c for c in sorted_contacts if int(c["Rank"]) >= 70]
    tier2 = [c for c in sorted_contacts if 40 <= int(c["Rank"]) < 70]
    tier3 = [c for c in sorted_contacts if int(c["Rank"]) < 40]

    wb = Workbook()

    # ── Sheet 1: Dashboard ──
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    build_dashboard(ws_dash, contacts)

    # ── Sheet 2: All Contacts ──
    ws_all = wb.create_sheet("All Contacts")
    add_headers(ws_all, CRM_COLUMNS, NAVY_FILL)
    add_contact_rows(ws_all, sorted_contacts)
    ws_all.freeze_panes = 'A2'
    ws_all.auto_filter.ref = f'A1:X{len(sorted_contacts)+1}'
    ws_all.row_dimensions[1].height = 30

    # ── Sheet 3: Tier 1 ──
    ws_t1 = wb.create_sheet("Tier 1 — Immediate")
    add_headers(ws_t1, CRM_COLUMNS, GREEN_HEADER)
    add_contact_rows(ws_t1, tier1)
    ws_t1.freeze_panes = 'A2'
    ws_t1.auto_filter.ref = f'A1:X{len(tier1)+1}'
    ws_t1.row_dimensions[1].height = 30

    # ── Sheet 4: Tier 2 ──
    ws_t2 = wb.create_sheet("Tier 2 — Next Wave")
    add_headers(ws_t2, CRM_COLUMNS, GOLD_HEADER)
    add_contact_rows(ws_t2, tier2)
    ws_t2.freeze_panes = 'A2'
    ws_t2.auto_filter.ref = f'A1:X{len(tier2)+1}'
    ws_t2.row_dimensions[1].height = 30

    # ── Sheet 5: Tier 3 ──
    ws_t3 = wb.create_sheet("Tier 3 — Long Game")
    add_headers(ws_t3, CRM_COLUMNS, GRAY_HEADER)
    add_contact_rows(ws_t3, tier3)
    ws_t3.freeze_panes = 'A2'
    ws_t3.auto_filter.ref = f'A1:X{len(tier3)+1}'
    ws_t3.row_dimensions[1].height = 30

    # ── Save ──
    output = os.path.join(BASE, "crm", "tastyr-iq-crm-workbook.xlsx")
    wb.save(output)
    print(f"CRM Workbook saved: {output}")
    print(f"  Dashboard sheet: summary stats, segment breakdown, chains, ranking formula")
    print(f"  All Contacts: {len(sorted_contacts)} rows with filters and freeze panes")
    print(f"  Tier 1: {len(tier1)} contacts (rank 70-100)")
    print(f"  Tier 2: {len(tier2)} contacts (rank 40-69)")
    print(f"  Tier 3: {len(tier3)} contacts (rank 1-39)")

if __name__ == "__main__":
    build_xlsx()
