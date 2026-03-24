#!/usr/bin/env python3
"""
FORK Financial Model Generator
Generates 3 Excel workbook variations (Conservative, Base, Aggressive)
with full revenue projections, cost structures, unit economics,
funding/runway analysis, and assumption-driven formulas.
"""

import os
import math
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palettes per variation
# ---------------------------------------------------------------------------
PALETTES = {
    "conservative": {
        "header_bg": "16A34A",
        "header_font": "FFFFFF",
        "accent": "22C55E",
        "alt_row": "F0FDF4",
        "title_font": "15803D",
        "border": "BBF7D0",
        "kpi_bg": "DCFCE7",
    },
    "base": {
        "header_bg": "0A1A2F",
        "header_font": "FFFFFF",
        "accent": "D4A534",
        "alt_row": "F0F4FF",
        "title_font": "0A1A2F",
        "border": "CBD5E1",
        "kpi_bg": "FEF9C3",
    },
    "aggressive": {
        "header_bg": "F97316",
        "header_font": "FFFFFF",
        "accent": "FB923C",
        "alt_row": "FFF7ED",
        "title_font": "C2410C",
        "border": "FED7AA",
        "kpi_bg": "FFEDD5",
    },
}

CURRENCY_FMT = '#,##0'
CURRENCY_FMT_DEC = '#,##0.00'
DOLLAR_FMT = '$#,##0'
DOLLAR_FMT_DEC = '$#,##0.00'
PCT_FMT = '0.0%'
PCT_FMT_INT = '0%'

MONTHS_2026 = [
    "Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26",
    "Jul-26", "Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26",
]
QUARTERS_2027 = ["Q1-27", "Q2-27", "Q3-27", "Q4-27"]
QUARTERS_2028 = ["Q1-28", "Q2-28", "Q3-28", "Q4-28"]
QUARTERS_2029 = ["Q1-29", "Q2-29", "Q3-29", "Q4-29"]


# ===================================================================
# Helper: styling utilities
# ===================================================================

def _header_fill(palette):
    return PatternFill(start_color=palette["header_bg"],
                       end_color=palette["header_bg"], fill_type="solid")

def _header_font(palette, size=11):
    return Font(name="Calibri", bold=True, color=palette["header_font"], size=size)

def _accent_fill(palette):
    return PatternFill(start_color=palette["accent"],
                       end_color=palette["accent"], fill_type="solid")

def _alt_fill(palette):
    return PatternFill(start_color=palette["alt_row"],
                       end_color=palette["alt_row"], fill_type="solid")

def _kpi_fill(palette):
    return PatternFill(start_color=palette["kpi_bg"],
                       end_color=palette["kpi_bg"], fill_type="solid")

def _title_font(palette, size=14):
    return Font(name="Calibri", bold=True, color=palette["title_font"], size=size)

def _bold():
    return Font(name="Calibri", bold=True, size=11)

def _normal():
    return Font(name="Calibri", size=11)

def _thin_border(palette):
    side = Side(style="thin", color=palette["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def _apply_header_row(ws, row, max_col, palette):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _header_fill(palette)
        cell.font = _header_font(palette)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border(palette)

def _apply_data_rows(ws, start_row, end_row, max_col, palette, fmt_map=None):
    """Apply alternating rows + borders + number formats.
    fmt_map: dict mapping 1-based row offset (relative to start_row) -> format string,
             or dict mapping col index -> format string.
    """
    for r in range(start_row, end_row + 1):
        is_alt = (r - start_row) % 2 == 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if is_alt:
                cell.fill = _alt_fill(palette)
            cell.border = _thin_border(palette)
            cell.font = _normal()
            cell.alignment = Alignment(horizontal="center", vertical="center")

def _set_col_widths(ws, widths):
    """widths: list of (col_letter_or_index, width)"""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def _auto_col_widths(ws, min_width=12, max_width=22):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

def _write_title(ws, row, title, palette, merge_end=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _title_font(palette, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _format_currency_col(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = DOLLAR_FMT

def _format_pct_col(ws, col, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = PCT_FMT


# ===================================================================
# Revenue / cost computation engine
# ===================================================================

def compute_monthly_2026(assumptions):
    """Return list of 12 monthly dicts with all line items."""
    a = assumptions
    months = []
    total_users = 0
    active_restaurants = 0

    for m in range(1, 13):  # 1 = Jan, 12 = Dec
        d = {"month": m}

        # -- B2C users --
        if m < a["b2c_start_month"]:
            total_users = 0
        elif m == a["b2c_start_month"]:
            total_users = a["b2c_initial_users"]
        else:
            total_users = int(total_users * (1 + a["b2c_mom_growth"]))
        d["total_users"] = total_users
        d["premium_users"] = int(total_users * a["b2c_premium_conv"])
        d["b2c_revenue"] = round(d["premium_users"] * a["b2c_price"], 2)

        # -- B2B restaurants --
        if m < a["b2b_start_month"]:
            active_restaurants = 0
        elif m == a["b2b_start_month"]:
            active_restaurants = a["b2b_initial_restaurants"]
        else:
            active_restaurants += a["b2b_monthly_add"]
        d["active_restaurants"] = active_restaurants
        d["b2b_revenue"] = round(active_restaurants * a["b2b_price"], 2)

        # -- Promoted dishes (starts in specified month/year) --
        if m >= a.get("promoted_start_month_2026", 99):
            base_promoted = a.get("promoted_base_monthly", 500)
            # scale with restaurant count
            d["promoted_revenue"] = round(base_promoted * max(1, active_restaurants / 5), 2)
        else:
            d["promoted_revenue"] = 0

        # Data licensing doesn't start in 2026 for any variation
        d["data_revenue"] = 0

        d["total_revenue"] = round(
            d["b2c_revenue"] + d["b2b_revenue"] + d["promoted_revenue"] + d["data_revenue"], 2
        )

        # -- Costs --
        d["engineering"] = a["eng_cost_monthly"]
        d["cloud"] = round(a["cloud_base"] + total_users * a["cloud_per_user"], 2)
        d["marketing"] = round(a["marketing_base"] + (a["marketing_growth_monthly"] * max(0, m - 3)), 2)
        d["legal"] = a["legal_monthly"]
        d["operations"] = a["ops_monthly"]
        d["total_cost"] = round(
            d["engineering"] + d["cloud"] + d["marketing"] + d["legal"] + d["operations"], 2
        )
        d["net"] = round(d["total_revenue"] - d["total_cost"], 2)

        months.append(d)
    return months


def compute_quarterly(assumptions, year, prev_end_users=0, prev_end_restaurants=0):
    """Return list of 4 quarterly dicts."""
    a = assumptions
    quarters = []
    total_users = prev_end_users
    active_restaurants = prev_end_restaurants

    year_offset = year - 2027  # 0, 1, 2

    for q in range(1, 5):
        d = {"quarter": q, "year": year}

        # Grow users for 3 months
        for _ in range(3):
            total_users = int(total_users * (1 + a["b2c_mom_growth"]))
        d["total_users"] = total_users
        d["premium_users"] = int(total_users * a["b2c_premium_conv"])
        # quarterly revenue = 3 months of premium
        d["b2c_revenue"] = round(d["premium_users"] * a["b2c_price"] * 3, 2)

        # Restaurants
        active_restaurants += a["b2b_monthly_add"] * 3
        d["active_restaurants"] = active_restaurants
        d["b2b_revenue"] = round(active_restaurants * a["b2b_price"] * 3, 2)

        # Data licensing
        quarter_global = (year - 2027) * 4 + q  # 1-based across years
        data_start_q = a.get("data_start_quarter_global", 3)  # e.g. Q3 2027 = 3
        if quarter_global >= data_start_q:
            periods_since = quarter_global - data_start_q
            d["data_revenue"] = round(
                a["data_initial_quarterly"] * (1 + a.get("data_growth_rate", 0.25)) ** periods_since, 2
            )
        else:
            d["data_revenue"] = 0

        # Promoted dishes
        promoted_start_q = a.get("promoted_start_quarter_global", 1)
        if quarter_global >= promoted_start_q:
            periods_since = quarter_global - promoted_start_q
            d["promoted_revenue"] = round(
                a.get("promoted_base_monthly", 500) * 3 *
                (1 + a.get("promoted_growth_rate", 0.15)) ** periods_since, 2
            )
        else:
            d["promoted_revenue"] = 0

        d["total_revenue"] = round(
            d["b2c_revenue"] + d["b2b_revenue"] + d["data_revenue"] + d["promoted_revenue"], 2
        )

        # Costs (quarterly)
        eng_multiplier = 1 + year_offset * a.get("eng_annual_growth", 0.5)
        d["engineering"] = round(a["eng_cost_monthly"] * eng_multiplier * 3, 2)
        d["cloud"] = round((a["cloud_base"] + total_users * a["cloud_per_user"]) * 3, 2)
        mkt_multiplier = 1 + year_offset * a.get("mkt_annual_growth", 0.6)
        d["marketing"] = round(a["marketing_base"] * mkt_multiplier * 3, 2)
        d["legal"] = round(a["legal_monthly"] * 3, 2)
        d["operations"] = round(a["ops_monthly"] * (1 + year_offset * 0.3) * 3, 2)
        d["total_cost"] = round(
            d["engineering"] + d["cloud"] + d["marketing"] + d["legal"] + d["operations"], 2
        )
        d["net"] = round(d["total_revenue"] - d["total_cost"], 2)

        quarters.append(d)

    return quarters, total_users, active_restaurants


# ===================================================================
# Assumptions for each variation
# ===================================================================

CONSERVATIVE = {
    "label": "Conservative",
    "tag": "v1-conservative",
    # B2C
    "b2c_start_month": 3,
    "b2c_initial_users": 100,
    "b2c_mom_growth": 0.15,
    "b2c_premium_conv": 0.05,
    "b2c_price": 4.99,
    # B2B
    "b2b_start_month": 4,
    "b2b_initial_restaurants": 3,
    "b2b_monthly_add": 2,
    "b2b_price": 500,
    # Data licensing
    "data_start_quarter_global": 7,  # Q3 2027 => global quarter 3, but across years: Q3-2027=3
    "data_initial_quarterly": 5000,
    "data_growth_rate": 0.25,
    # Promoted dishes
    "promoted_start_month_2026": 99,  # not in 2026
    "promoted_start_quarter_global": 1,  # Q1 2027
    "promoted_base_monthly": 500,
    "promoted_growth_rate": 0.12,
    # Costs
    "eng_cost_monthly": 15000,
    "eng_annual_growth": 0.40,
    "cloud_base": 500,
    "cloud_per_user": 0.10,
    "marketing_base": 3000,
    "marketing_growth_monthly": 200,
    "mkt_annual_growth": 0.50,
    "legal_monthly": 2000,
    "ops_monthly": 2000,
    # Funding
    "raise_amount": 750000,
    "pre_money": 3000000,
    # Unit economics
    "cac_b2c": 8,
    "cac_b2b": 2500,
    "ltv_b2c_months": 14,
    "ltv_b2b_months": 24,
    "gross_margin_b2c": 0.85,
    "gross_margin_b2b": 0.75,
    # Targets
    "target_arr_2026": 120000,
    "target_arr_2027": 1200000,
    "target_arr_2029": 8000000,
}

BASE_CASE = {
    "label": "Base Case",
    "tag": "v2-base",
    "b2c_start_month": 3,
    "b2c_initial_users": 100,
    "b2c_mom_growth": 0.20,
    "b2c_premium_conv": 0.08,
    "b2c_price": 4.99,
    "b2b_start_month": 3,
    "b2b_initial_restaurants": 5,
    "b2b_monthly_add": 3,
    "b2b_price": 500,
    "data_start_quarter_global": 6,  # Q2 2027
    "data_initial_quarterly": 8000,
    "data_growth_rate": 0.30,
    "promoted_start_month_2026": 99,
    "promoted_start_quarter_global": 1,
    "promoted_base_monthly": 800,
    "promoted_growth_rate": 0.18,
    "eng_cost_monthly": 18000,
    "eng_annual_growth": 0.55,
    "cloud_base": 600,
    "cloud_per_user": 0.08,
    "marketing_base": 5000,
    "marketing_growth_monthly": 350,
    "mkt_annual_growth": 0.65,
    "legal_monthly": 2000,
    "ops_monthly": 2500,
    "raise_amount": 1000000,
    "pre_money": 4500000,
    "cac_b2c": 6,
    "cac_b2b": 2000,
    "ltv_b2c_months": 16,
    "ltv_b2b_months": 30,
    "gross_margin_b2c": 0.87,
    "gross_margin_b2b": 0.78,
    "target_arr_2026": 120000,
    "target_arr_2027": 1200000,
    "target_arr_2029": 18000000,
}

AGGRESSIVE = {
    "label": "Aggressive / Upside",
    "tag": "v3-aggressive",
    "b2c_start_month": 2,
    "b2c_initial_users": 150,
    "b2c_mom_growth": 0.25,
    "b2c_premium_conv": 0.12,
    "b2c_price": 4.99,
    "b2b_start_month": 2,
    "b2b_initial_restaurants": 8,
    "b2b_monthly_add": 5,
    "b2b_price": 500,
    "data_start_quarter_global": 5,  # Q1 2027
    "data_initial_quarterly": 10000,
    "data_growth_rate": 0.35,
    "promoted_start_month_2026": 6,
    "promoted_start_quarter_global": 1,
    "promoted_base_monthly": 1200,
    "promoted_growth_rate": 0.22,
    "eng_cost_monthly": 22000,
    "eng_annual_growth": 0.70,
    "cloud_base": 800,
    "cloud_per_user": 0.06,
    "marketing_base": 7000,
    "marketing_growth_monthly": 500,
    "mkt_annual_growth": 0.80,
    "legal_monthly": 2500,
    "ops_monthly": 3000,
    "raise_amount": 1500000,
    "pre_money": 6000000,
    "cac_b2c": 5,
    "cac_b2b": 1800,
    "ltv_b2c_months": 18,
    "ltv_b2b_months": 36,
    "gross_margin_b2c": 0.90,
    "gross_margin_b2b": 0.80,
    "target_arr_2026": 120000,
    "target_arr_2027": 2500000,
    "target_arr_2029": 30000000,
    # Aggressive-specific extras
    "viral_coefficient": 1.3,
    "engineers_eoy_2026": 4,
    "engineers_2028": 12,
}


# ===================================================================
# Sheet builders
# ===================================================================

def build_assumptions_sheet(wb, a, palette):
    ws = wb.create_sheet("Assumptions")
    _write_title(ws, 1, "FORK Financial Model -- Key Assumptions", palette, merge_end=4)

    headers = ["Category", "Parameter", "Value", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _apply_header_row(ws, 3, 4, palette)

    rows = [
        ("B2C", "Launch month (2026)", a["b2c_start_month"], f"Month {a['b2c_start_month']} = {MONTHS_2026[a['b2c_start_month']-1]}"),
        ("B2C", "Initial users at launch", a["b2c_initial_users"], "Free users"),
        ("B2C", "Month-over-month growth", a["b2c_mom_growth"], "Organic + paid"),
        ("B2C", "Premium conversion rate", a["b2c_premium_conv"], "Free-to-paid"),
        ("B2C", "Subscription price", a["b2c_price"], "$/month"),
        ("", "", "", ""),
        ("B2B", "Launch month (2026)", a["b2b_start_month"], f"Month {a['b2b_start_month']}"),
        ("B2B", "Initial restaurant partners", a["b2b_initial_restaurants"], "At launch"),
        ("B2B", "New restaurants / month", a["b2b_monthly_add"], "Sales pipeline"),
        ("B2B", "Monthly SaaS fee", a["b2b_price"], "$/restaurant/month"),
        ("", "", "", ""),
        ("Data Licensing", "Start quarter (global)", a.get("data_start_quarter_global", "N/A"), "Quarters from Q1-2027"),
        ("Data Licensing", "Initial quarterly revenue", a.get("data_initial_quarterly", 0), "$"),
        ("Data Licensing", "Quarterly growth rate", a.get("data_growth_rate", 0), "Compound"),
        ("", "", "", ""),
        ("Promoted Dishes", "Base monthly revenue", a.get("promoted_base_monthly", 0), "$"),
        ("Promoted Dishes", "Growth rate per quarter", a.get("promoted_growth_rate", 0), ""),
        ("", "", "", ""),
        ("Engineering", "Monthly cost (initial)", a["eng_cost_monthly"], "$"),
        ("Engineering", "Annual cost growth", a.get("eng_annual_growth", 0.5), "Hiring ramp"),
        ("Cloud/Infra", "Base monthly", a["cloud_base"], "$"),
        ("Cloud/Infra", "Per-user increment", a["cloud_per_user"], "$/user/month"),
        ("Marketing", "Base monthly", a["marketing_base"], "$"),
        ("Marketing", "Annual growth multiplier", a.get("mkt_annual_growth", 0.5), ""),
        ("Legal/IP", "Monthly", a["legal_monthly"], "Patent prosecution"),
        ("Operations", "Monthly", a["ops_monthly"], "$"),
        ("", "", "", ""),
        ("Funding", "Raise amount", a["raise_amount"], "$"),
        ("Funding", "Pre-money valuation", a["pre_money"], "$"),
        ("Funding", "Post-money valuation", a["raise_amount"] + a["pre_money"], "$"),
        ("Funding", "Dilution", round(a["raise_amount"] / (a["raise_amount"] + a["pre_money"]) * 100, 1), "%"),
        ("", "", "", ""),
        ("Unit Economics", "CAC (B2C)", a["cac_b2c"], "$"),
        ("Unit Economics", "CAC (B2B)", a["cac_b2b"], "$"),
        ("Unit Economics", "Avg B2C lifetime (months)", a["ltv_b2c_months"], ""),
        ("Unit Economics", "Avg B2B lifetime (months)", a["ltv_b2b_months"], ""),
        ("Unit Economics", "Gross margin B2C", a["gross_margin_b2c"], ""),
        ("Unit Economics", "Gross margin B2B", a["gross_margin_b2b"], ""),
    ]

    for i, (cat, param, val, note) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=cat).font = _bold() if cat else _normal()
        ws.cell(row=r, column=2, value=param)
        cell_val = ws.cell(row=r, column=3, value=val)
        if isinstance(val, float) and val < 1:
            cell_val.number_format = PCT_FMT
        elif isinstance(val, (int, float)) and val >= 100:
            cell_val.number_format = DOLLAR_FMT
        ws.cell(row=r, column=4, value=note)

    end_row = 4 + len(rows) - 1
    _apply_data_rows(ws, 4, end_row, 4, palette)
    _auto_col_widths(ws, min_width=14, max_width=40)
    ws.freeze_panes = "A4"
    ws.sheet_properties.pageSetUpPr = None  # reset


def build_revenue_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029):
    ws = wb.create_sheet("Revenue Projections")
    _write_title(ws, 1, f"FORK Revenue Projections -- {a['label']}", palette, merge_end=8)

    # ---- 2026 Monthly ----
    ws.cell(row=3, column=1, value="2026 Monthly Breakdown").font = _bold()
    headers_2026 = ["Metric"] + MONTHS_2026 + ["2026 Total"]
    for c, h in enumerate(headers_2026, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_row(ws, 4, len(headers_2026), palette)

    metric_keys = [
        ("Total Users", "total_users", CURRENCY_FMT),
        ("Premium Users", "premium_users", CURRENCY_FMT),
        ("B2C Revenue", "b2c_revenue", DOLLAR_FMT),
        ("Active Restaurants", "active_restaurants", CURRENCY_FMT),
        ("B2B Revenue", "b2b_revenue", DOLLAR_FMT),
        ("Promoted Dishes Rev", "promoted_revenue", DOLLAR_FMT),
        ("Data Licensing Rev", "data_revenue", DOLLAR_FMT),
        ("TOTAL REVENUE", "total_revenue", DOLLAR_FMT),
    ]

    for mi, (label, key, fmt) in enumerate(metric_keys):
        r = 5 + mi
        ws.cell(row=r, column=1, value=label).font = _bold() if key == "total_revenue" else _normal()
        yearly_total = 0
        for m_idx, md in enumerate(monthly_data):
            val = md[key]
            cell = ws.cell(row=r, column=2 + m_idx, value=val)
            cell.number_format = fmt
            if key in ("b2c_revenue", "b2b_revenue", "promoted_revenue", "data_revenue", "total_revenue"):
                yearly_total += val
            else:
                yearly_total = val  # for counts, show last month
        # Total column
        total_cell = ws.cell(row=r, column=14, value=yearly_total)
        total_cell.number_format = fmt
        total_cell.font = _bold()

    end_2026 = 5 + len(metric_keys) - 1
    _apply_data_rows(ws, 5, end_2026, 14, palette)

    # Highlight total revenue row
    for c in range(1, 15):
        ws.cell(row=end_2026, column=c).fill = _kpi_fill(palette)
        ws.cell(row=end_2026, column=c).font = _bold()

    # ---- 2027-2029 Quarterly ----
    q_start_row = end_2026 + 3
    ws.cell(row=q_start_row - 1, column=1, value="2027-2029 Quarterly Breakdown").font = _bold()

    q_headers = ["Metric"] + QUARTERS_2027 + QUARTERS_2028 + QUARTERS_2029
    for c, h in enumerate(q_headers, 1):
        ws.cell(row=q_start_row, column=c, value=h)
    _apply_header_row(ws, q_start_row, len(q_headers), palette)

    all_q = q2027 + q2028 + q2029  # 12 quarters

    for mi, (label, key, fmt) in enumerate(metric_keys):
        r = q_start_row + 1 + mi
        ws.cell(row=r, column=1, value=label).font = _bold() if key == "total_revenue" else _normal()
        for qi, qd in enumerate(all_q):
            cell = ws.cell(row=r, column=2 + qi, value=qd[key])
            cell.number_format = fmt

    end_q = q_start_row + 1 + len(metric_keys) - 1
    _apply_data_rows(ws, q_start_row + 1, end_q, 1 + len(all_q), palette)

    # Highlight total revenue row
    for c in range(1, 2 + len(all_q)):
        ws.cell(row=end_q, column=c).fill = _kpi_fill(palette)
        ws.cell(row=end_q, column=c).font = _bold()

    # ---- ARR Summary ----
    arr_row = end_q + 3
    ws.cell(row=arr_row, column=1, value="Annual Run Rate (ARR)").font = _bold()
    years_labels = ["2026", "2027", "2028", "2029"]
    rev_2026_total = sum(md["total_revenue"] for md in monthly_data)
    rev_2027_total = sum(qd["total_revenue"] for qd in q2027)
    rev_2028_total = sum(qd["total_revenue"] for qd in q2028)
    rev_2029_total = sum(qd["total_revenue"] for qd in q2029)
    # ARR = last period annualized
    arr_2026 = monthly_data[-1]["total_revenue"] * 12
    arr_2027 = q2027[-1]["total_revenue"] * 4
    arr_2028 = q2028[-1]["total_revenue"] * 4
    arr_2029 = q2029[-1]["total_revenue"] * 4

    for c, lbl in enumerate(years_labels, 2):
        ws.cell(row=arr_row, column=c, value=lbl).font = _bold()
    ws.cell(row=arr_row + 1, column=1, value="Total Revenue")
    for c, val in enumerate([rev_2026_total, rev_2027_total, rev_2028_total, rev_2029_total], 2):
        cell = ws.cell(row=arr_row + 1, column=c, value=round(val))
        cell.number_format = DOLLAR_FMT
    ws.cell(row=arr_row + 2, column=1, value="ARR (exit run-rate)")
    for c, val in enumerate([arr_2026, arr_2027, arr_2028, arr_2029], 2):
        cell = ws.cell(row=arr_row + 2, column=c, value=round(val))
        cell.number_format = DOLLAR_FMT
        cell.font = _bold()
        cell.fill = _kpi_fill(palette)

    _apply_header_row(ws, arr_row, 5, palette)

    # ---- Revenue Bar Chart ----
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Quarterly Revenue by Stream (2027-2029)"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Quarter"
    chart.width = 28
    chart.height = 14

    # Data for chart: rows for B2C, B2B, Data, Promoted across quarters
    chart_start = end_q + 6
    chart_labels = QUARTERS_2027 + QUARTERS_2028 + QUARTERS_2029
    ws.cell(row=chart_start, column=1, value="Quarter")
    ws.cell(row=chart_start + 1, column=1, value="B2C Revenue")
    ws.cell(row=chart_start + 2, column=1, value="B2B Revenue")
    ws.cell(row=chart_start + 3, column=1, value="Data Licensing")
    ws.cell(row=chart_start + 4, column=1, value="Promoted Dishes")
    for qi, qd in enumerate(all_q):
        col = 2 + qi
        ws.cell(row=chart_start, column=col, value=chart_labels[qi])
        ws.cell(row=chart_start + 1, column=col, value=round(qd["b2c_revenue"]))
        ws.cell(row=chart_start + 2, column=col, value=round(qd["b2b_revenue"]))
        ws.cell(row=chart_start + 3, column=col, value=round(qd["data_revenue"]))
        ws.cell(row=chart_start + 4, column=col, value=round(qd["promoted_revenue"]))

    cats = Reference(ws, min_col=2, max_col=1 + len(all_q), min_row=chart_start)
    for series_row in range(chart_start + 1, chart_start + 5):
        data = Reference(ws, min_col=2, max_col=1 + len(all_q), min_row=series_row)
        chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].tx = SeriesLabel(v="B2C Revenue")
    chart.series[1].tx = SeriesLabel(v="B2B Revenue")
    chart.series[2].tx = SeriesLabel(v="Data Licensing")
    chart.series[3].tx = SeriesLabel(v="Promoted Dishes")
    chart.shape = 4
    ws.add_chart(chart, f"A{chart_start + 6}")

    _auto_col_widths(ws, min_width=13, max_width=18)
    ws.freeze_panes = "B5"

    return {
        "rev_2026": rev_2026_total,
        "rev_2027": rev_2027_total,
        "rev_2028": rev_2028_total,
        "rev_2029": rev_2029_total,
        "arr_2026": arr_2026,
        "arr_2027": arr_2027,
        "arr_2028": arr_2028,
        "arr_2029": arr_2029,
    }


def build_cost_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029):
    ws = wb.create_sheet("Cost Structure")
    _write_title(ws, 1, f"FORK Cost Structure -- {a['label']}", palette, merge_end=8)

    # 2026 Monthly
    ws.cell(row=3, column=1, value="2026 Monthly Costs").font = _bold()
    headers = ["Cost Category"] + MONTHS_2026 + ["2026 Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_row(ws, 4, len(headers), palette)

    cost_keys = [
        ("Engineering", "engineering"),
        ("Cloud / Infrastructure", "cloud"),
        ("Marketing", "marketing"),
        ("Legal / IP", "legal"),
        ("Operations", "operations"),
        ("TOTAL COSTS", "total_cost"),
        ("NET (Revenue - Costs)", "net"),
    ]

    for ci, (label, key) in enumerate(cost_keys):
        r = 5 + ci
        ws.cell(row=r, column=1, value=label).font = _bold() if "TOTAL" in label or "NET" in label else _normal()
        yearly = 0
        for mi, md in enumerate(monthly_data):
            val = md[key]
            cell = ws.cell(row=r, column=2 + mi, value=round(val))
            cell.number_format = DOLLAR_FMT
            yearly += val
        total_cell = ws.cell(row=r, column=14, value=round(yearly))
        total_cell.number_format = DOLLAR_FMT
        total_cell.font = _bold()

    end_2026 = 5 + len(cost_keys) - 1
    _apply_data_rows(ws, 5, end_2026, 14, palette)

    # Highlight net row
    for c in range(1, 15):
        ws.cell(row=end_2026, column=c).fill = _kpi_fill(palette)

    # 2027-2029 quarterly
    q_start = end_2026 + 3
    ws.cell(row=q_start - 1, column=1, value="2027-2029 Quarterly Costs").font = _bold()
    q_headers = ["Cost Category"] + QUARTERS_2027 + QUARTERS_2028 + QUARTERS_2029
    for c, h in enumerate(q_headers, 1):
        ws.cell(row=q_start, column=c, value=h)
    _apply_header_row(ws, q_start, len(q_headers), palette)

    all_q = q2027 + q2028 + q2029
    for ci, (label, key) in enumerate(cost_keys):
        r = q_start + 1 + ci
        ws.cell(row=r, column=1, value=label).font = _bold() if "TOTAL" in label or "NET" in label else _normal()
        for qi, qd in enumerate(all_q):
            cell = ws.cell(row=r, column=2 + qi, value=round(qd[key]))
            cell.number_format = DOLLAR_FMT

    end_q = q_start + 1 + len(cost_keys) - 1
    _apply_data_rows(ws, q_start + 1, end_q, 1 + len(all_q), palette)
    for c in range(1, 2 + len(all_q)):
        ws.cell(row=end_q, column=c).fill = _kpi_fill(palette)

    _auto_col_widths(ws, min_width=13, max_width=18)
    ws.freeze_panes = "B5"


def build_unit_economics_sheet(wb, a, palette):
    ws = wb.create_sheet("Unit Economics")
    _write_title(ws, 1, f"FORK Unit Economics -- {a['label']}", palette, merge_end=5)

    headers = ["Metric", "B2C (Consumer)", "B2B (Restaurant SaaS)", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _apply_header_row(ws, 3, 4, palette)

    ltv_b2c = round(a["b2c_price"] * a["ltv_b2c_months"] * a["gross_margin_b2c"], 2)
    ltv_b2b = round(a["b2b_price"] * a["ltv_b2b_months"] * a["gross_margin_b2b"], 2)
    ltv_cac_b2c = round(ltv_b2c / a["cac_b2c"], 1) if a["cac_b2c"] else 0
    ltv_cac_b2b = round(ltv_b2b / a["cac_b2b"], 1) if a["cac_b2b"] else 0
    payback_b2c = round(a["cac_b2c"] / (a["b2c_price"] * a["gross_margin_b2c"]), 1)
    payback_b2b = round(a["cac_b2b"] / (a["b2b_price"] * a["gross_margin_b2b"]), 1)

    rows = [
        ("Monthly ARPU", f"${a['b2c_price']}", f"${a['b2b_price']}", "Revenue per user/account"),
        ("Customer Acquisition Cost (CAC)", f"${a['cac_b2c']}", f"${a['cac_b2b']:,}", "Blended cost"),
        ("Avg Lifetime (months)", a["ltv_b2c_months"], a["ltv_b2b_months"], "Before churn"),
        ("Gross Margin", f"{a['gross_margin_b2c']:.0%}", f"{a['gross_margin_b2b']:.0%}", "After COGS"),
        ("Lifetime Value (LTV)", f"${ltv_b2c:,.0f}", f"${ltv_b2b:,.0f}", "ARPU x Lifetime x Margin"),
        ("LTV / CAC Ratio", f"{ltv_cac_b2c}x", f"{ltv_cac_b2b}x", "Target: >3x"),
        ("Payback Period (months)", f"{payback_b2c}", f"{payback_b2b}", "CAC / (ARPU x Margin)"),
        ("", "", "", ""),
        ("Revenue per Restaurant", "--", "$6,000/yr", "SaaS subscription"),
        ("Expansion Revenue Potential", "Upsell tiers", "Data add-ons", "Future ARPU growth"),
    ]

    for i, (metric, b2c, b2b, note) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=metric).font = _bold() if "LTV / CAC" in metric else _normal()
        ws.cell(row=r, column=2, value=b2c)
        ws.cell(row=r, column=3, value=b2b)
        ws.cell(row=r, column=4, value=note)

    _apply_data_rows(ws, 4, 4 + len(rows) - 1, 4, palette)

    # Highlight LTV/CAC row
    ltv_row = 4 + 5  # 0-indexed position of LTV/CAC
    for c in range(1, 5):
        ws.cell(row=ltv_row, column=c).fill = _kpi_fill(palette)
        ws.cell(row=ltv_row, column=c).font = _bold()

    _auto_col_widths(ws, min_width=18, max_width=35)
    ws.freeze_panes = "A4"


def build_funding_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029, rev_summary):
    ws = wb.create_sheet("Funding & Runway")
    _write_title(ws, 1, f"FORK Funding & Runway -- {a['label']}", palette, merge_end=6)

    # Funding summary
    ws.cell(row=3, column=1, value="Funding Round Summary").font = _bold()
    fund_items = [
        ("Raise Amount", a["raise_amount"], DOLLAR_FMT),
        ("Pre-Money Valuation", a["pre_money"], DOLLAR_FMT),
        ("Post-Money Valuation", a["raise_amount"] + a["pre_money"], DOLLAR_FMT),
        ("Investor Ownership", round(a["raise_amount"] / (a["raise_amount"] + a["pre_money"]) * 100, 1), "0.0\"%\""),
        ("", "", ""),
    ]
    for i, (lbl, val, fmt) in enumerate(fund_items):
        r = 4 + i
        ws.cell(row=r, column=1, value=lbl).font = _bold()
        cell = ws.cell(row=r, column=2, value=val)
        if fmt:
            cell.number_format = fmt

    # Monthly cash position (2026)
    cash_start_row = 4 + len(fund_items) + 1
    ws.cell(row=cash_start_row - 1, column=1, value="2026 Monthly Cash Position").font = _bold()
    cash_headers = ["Month", "Revenue", "Costs", "Net", "Cash Balance"]
    for c, h in enumerate(cash_headers, 1):
        ws.cell(row=cash_start_row, column=c, value=h)
    _apply_header_row(ws, cash_start_row, 5, palette)

    cash = a["raise_amount"]
    cash_data_rows = []
    for mi, md in enumerate(monthly_data):
        r = cash_start_row + 1 + mi
        cash += md["net"]
        ws.cell(row=r, column=1, value=MONTHS_2026[mi])
        ws.cell(row=r, column=2, value=round(md["total_revenue"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=3, value=round(md["total_cost"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=4, value=round(md["net"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=5, value=round(cash)).number_format = DOLLAR_FMT
        cash_data_rows.append(round(cash))

    end_cash_2026 = cash_start_row + 12
    _apply_data_rows(ws, cash_start_row + 1, end_cash_2026, 5, palette)

    # Quarterly cash (2027-2029)
    qcash_start = end_cash_2026 + 2
    ws.cell(row=qcash_start - 1, column=1, value="2027-2029 Quarterly Cash Position").font = _bold()
    for c, h in enumerate(cash_headers, 1):
        ws.cell(row=qcash_start, column=c, value=h.replace("Month", "Quarter"))
    _apply_header_row(ws, qcash_start, 5, palette)

    all_q = q2027 + q2028 + q2029
    q_labels = QUARTERS_2027 + QUARTERS_2028 + QUARTERS_2029
    q_cash_rows = []
    for qi, qd in enumerate(all_q):
        r = qcash_start + 1 + qi
        cash += qd["net"]
        ws.cell(row=r, column=1, value=q_labels[qi])
        ws.cell(row=r, column=2, value=round(qd["total_revenue"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=3, value=round(qd["total_cost"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=4, value=round(qd["net"])).number_format = DOLLAR_FMT
        ws.cell(row=r, column=5, value=round(cash)).number_format = DOLLAR_FMT
        q_cash_rows.append(round(cash))

    end_qcash = qcash_start + len(all_q)
    _apply_data_rows(ws, qcash_start + 1, end_qcash, 5, palette)

    # Runway calculation
    rw_row = end_qcash + 3
    avg_monthly_burn = round(sum(md["total_cost"] for md in monthly_data) / 12)
    remaining_cash = cash_data_rows[-1]  # end of 2026
    runway_months = round(remaining_cash / avg_monthly_burn, 1) if avg_monthly_burn > 0 else 999

    ws.cell(row=rw_row, column=1, value="Runway Analysis").font = _bold()
    rw_items = [
        ("Average Monthly Burn (2026)", avg_monthly_burn, DOLLAR_FMT),
        ("Cash Remaining EOY 2026", remaining_cash, DOLLAR_FMT),
        ("Runway from EOY 2026 (months)", runway_months, "0.0"),
        ("Break-even Quarter (est.)", "", ""),
    ]
    # Find break-even
    be_label = "Not in forecast"
    for qi, qd in enumerate(all_q):
        if qd["net"] > 0:
            be_label = q_labels[qi]
            break
    rw_items[3] = ("Break-even Quarter (est.)", be_label, "@")

    for i, (lbl, val, fmt) in enumerate(rw_items):
        r = rw_row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _bold()
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.fill = _kpi_fill(palette)

    # Cash position line chart
    chart = LineChart()
    chart.title = "Cash Position Over Time"
    chart.y_axis.title = "Cash ($)"
    chart.x_axis.title = "Period"
    chart.width = 24
    chart.height = 12
    chart.style = 10

    chart_row = rw_row + 7
    all_cash = cash_data_rows + q_cash_rows
    all_labels = MONTHS_2026 + q_labels
    ws.cell(row=chart_row, column=1, value="Period")
    ws.cell(row=chart_row + 1, column=1, value="Cash Balance")
    for ci, (lbl, cv) in enumerate(zip(all_labels, all_cash)):
        ws.cell(row=chart_row, column=2 + ci, value=lbl)
        ws.cell(row=chart_row + 1, column=2 + ci, value=cv)

    cats = Reference(ws, min_col=2, max_col=1 + len(all_cash), min_row=chart_row)
    data = Reference(ws, min_col=2, max_col=1 + len(all_cash), min_row=chart_row + 1)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.series[0].tx = SeriesLabel(v="Cash Balance")
    ws.add_chart(chart, f"A{chart_row + 3}")

    _auto_col_widths(ws, min_width=14, max_width=22)
    ws.freeze_panes = f"A{cash_start_row + 1}"

    return {
        "avg_burn": avg_monthly_burn,
        "cash_eoy_2026": remaining_cash,
        "runway_months": runway_months,
        "break_even": be_label,
        "final_cash": round(cash),
    }


def build_summary_dashboard(wb, a, palette, rev_summary, funding_summary, monthly_data):
    ws = wb.create_sheet("Summary Dashboard")
    _write_title(ws, 1, f"FORK Financial Model -- {a['label']} Summary", palette, merge_end=6)
    ws.cell(row=2, column=1, value="AI-Powered Dish Intelligence Platform | Philadelphia Launch").font = Font(
        name="Calibri", italic=True, size=11, color="666666"
    )

    # Key metrics table
    ws.cell(row=4, column=1, value="KEY METRICS").font = _bold()
    kpi_headers = ["Metric", "2026", "2027", "2028", "2029"]
    for c, h in enumerate(kpi_headers, 1):
        ws.cell(row=5, column=c, value=h)
    _apply_header_row(ws, 5, 5, palette)

    kpi_rows = [
        ("Total Revenue", rev_summary["rev_2026"], rev_summary["rev_2027"],
         rev_summary["rev_2028"], rev_summary["rev_2029"]),
        ("ARR (exit run-rate)", rev_summary["arr_2026"], rev_summary["arr_2027"],
         rev_summary["arr_2028"], rev_summary["arr_2029"]),
        ("End Users", monthly_data[-1]["total_users"], "--", "--", "--"),
        ("Restaurant Partners", monthly_data[-1]["active_restaurants"], "--", "--", "--"),
    ]
    for i, (lbl, *vals) in enumerate(kpi_rows):
        r = 6 + i
        ws.cell(row=r, column=1, value=lbl).font = _bold()
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=v)
            if isinstance(v, (int, float)):
                cell.number_format = DOLLAR_FMT
            cell.fill = _kpi_fill(palette)

    _apply_data_rows(ws, 6, 6 + len(kpi_rows) - 1, 5, palette)

    # Funding summary
    fund_row = 6 + len(kpi_rows) + 2
    ws.cell(row=fund_row, column=1, value="FUNDING SUMMARY").font = _bold()
    fund_headers = ["Parameter", "Value"]
    for c, h in enumerate(fund_headers, 1):
        ws.cell(row=fund_row + 1, column=c, value=h)
    _apply_header_row(ws, fund_row + 1, 2, palette)

    post_money = a["raise_amount"] + a["pre_money"]
    fund_data = [
        ("Raise Target", a["raise_amount"]),
        ("Pre-Money Valuation", a["pre_money"]),
        ("Post-Money Valuation", post_money),
        ("Investor Ownership", f"{a['raise_amount'] / post_money * 100:.1f}%"),
        ("Avg Monthly Burn", funding_summary["avg_burn"]),
        ("Cash EOY 2026", funding_summary["cash_eoy_2026"]),
        ("Runway from EOY 2026", f"{funding_summary['runway_months']} months"),
        ("Est. Break-Even", funding_summary["break_even"]),
    ]
    for i, (lbl, val) in enumerate(fund_data):
        r = fund_row + 2 + i
        ws.cell(row=r, column=1, value=lbl).font = _bold()
        cell = ws.cell(row=r, column=2, value=val)
        if isinstance(val, (int, float)):
            cell.number_format = DOLLAR_FMT

    _apply_data_rows(ws, fund_row + 2, fund_row + 2 + len(fund_data) - 1, 2, palette)

    # Revenue streams description
    stream_row = fund_row + 2 + len(fund_data) + 2
    ws.cell(row=stream_row, column=1, value="REVENUE STREAMS").font = _bold()
    streams = [
        ("B2C Subscriptions", "$4.99/mo premium", "Dish ratings, personalized recommendations"),
        ("B2B Restaurant SaaS", "$500/mo per restaurant", "Dashboard, analytics, menu optimization"),
        ("Data Licensing", "Enterprise contracts", "Aggregated dish intelligence data"),
        ("Promoted Dishes", "CPC model", "Restaurants pay for dish visibility"),
    ]
    stream_headers = ["Stream", "Pricing", "Description"]
    for c, h in enumerate(stream_headers, 1):
        ws.cell(row=stream_row + 1, column=c, value=h)
    _apply_header_row(ws, stream_row + 1, 3, palette)
    for i, (s, p, d) in enumerate(streams):
        r = stream_row + 2 + i
        ws.cell(row=r, column=1, value=s).font = _bold()
        ws.cell(row=r, column=2, value=p)
        ws.cell(row=r, column=3, value=d)
    _apply_data_rows(ws, stream_row + 2, stream_row + 2 + len(streams) - 1, 3, palette)

    _auto_col_widths(ws, min_width=16, max_width=40)
    ws.freeze_panes = "A6"


def build_scenario_analysis_sheet(wb, palette, all_scenarios):
    """Side-by-side comparison of 3 scenarios (for Base Case workbook)."""
    ws = wb.create_sheet("Scenario Analysis")
    _write_title(ws, 1, "FORK Scenario Analysis -- 3 Scenarios Side-by-Side", palette, merge_end=6)

    headers = ["Metric", "Pessimistic (Conservative)", "Base Case", "Optimistic (Aggressive)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _apply_header_row(ws, 3, 4, palette)

    rows = [
        ("B2C MoM Growth", "15%", "20%", "25%"),
        ("Premium Conversion", "5%", "8%", "12%"),
        ("B2B Start Restaurants", "3", "5", "8"),
        ("B2B Monthly Adds", "2/mo", "3/mo", "5/mo"),
        ("", "", "", ""),
        ("2026 Total Revenue", None, None, None),
        ("2027 Total Revenue", None, None, None),
        ("2028 Total Revenue", None, None, None),
        ("2029 Total Revenue", None, None, None),
        ("", "", "", ""),
        ("2026 ARR", None, None, None),
        ("2027 ARR", None, None, None),
        ("2028 ARR", None, None, None),
        ("2029 ARR", None, None, None),
        ("", "", "", ""),
        ("Raise Amount", "$750K", "$1M", "$1.5M"),
        ("Pre-Money Valuation", "$3M", "$4.5M", "$6M"),
        ("Runway (from EOY 2026)", None, None, None),
        ("Est. Break-Even", None, None, None),
    ]

    # Fill computed values
    labels_rev = ["2026 Total Revenue", "2027 Total Revenue", "2028 Total Revenue", "2029 Total Revenue"]
    labels_arr = ["2026 ARR", "2027 ARR", "2028 ARR", "2029 ARR"]
    rev_keys = ["rev_2026", "rev_2027", "rev_2028", "rev_2029"]
    arr_keys = ["arr_2026", "arr_2027", "arr_2028", "arr_2029"]

    for i, row_data in enumerate(rows):
        r = 4 + i
        label = row_data[0]
        ws.cell(row=r, column=1, value=label).font = _bold() if label else _normal()

        if label in labels_rev:
            idx = labels_rev.index(label)
            for sc_i, sc in enumerate(all_scenarios):
                cell = ws.cell(row=r, column=2 + sc_i, value=round(sc["rev"][rev_keys[idx]]))
                cell.number_format = DOLLAR_FMT
        elif label in labels_arr:
            idx = labels_arr.index(label)
            for sc_i, sc in enumerate(all_scenarios):
                cell = ws.cell(row=r, column=2 + sc_i, value=round(sc["rev"][arr_keys[idx]]))
                cell.number_format = DOLLAR_FMT
                cell.fill = _kpi_fill(palette)
                cell.font = _bold()
        elif label == "Runway (from EOY 2026)":
            for sc_i, sc in enumerate(all_scenarios):
                ws.cell(row=r, column=2 + sc_i, value=f"{sc['fund']['runway_months']} months")
        elif label == "Est. Break-Even":
            for sc_i, sc in enumerate(all_scenarios):
                ws.cell(row=r, column=2 + sc_i, value=sc["fund"]["break_even"])
        else:
            for c_idx in range(1, 4):
                if row_data[c_idx] is not None:
                    ws.cell(row=r, column=1 + c_idx, value=row_data[c_idx])

    _apply_data_rows(ws, 4, 4 + len(rows) - 1, 4, palette)
    _auto_col_widths(ws, min_width=20, max_width=35)
    ws.freeze_panes = "A4"


def build_investor_returns_sheet(wb, a, palette, rev_summary):
    """Show ROI at different exit multiples for $500K and $1.5M investments."""
    ws = wb.create_sheet("Investor Returns")
    _write_title(ws, 1, f"FORK Investor Returns Analysis -- {a['label']}", palette, merge_end=7)

    post_money = a["raise_amount"] + a["pre_money"]
    investor_pct = a["raise_amount"] / post_money

    # Table 1: Exit multiples at different ARR points
    ws.cell(row=3, column=1, value="Exit Scenario Analysis").font = _bold()
    ws.cell(row=4, column=1, value=f"Based on {a['label']} ARR projections").font = Font(
        name="Calibri", italic=True, size=10, color="666666"
    )

    headers = ["Exit Year", "Projected ARR", "5x Revenue Multiple",
               "10x Revenue Multiple", "20x Revenue Multiple"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=6, column=c, value=h)
    _apply_header_row(ws, 6, 5, palette)

    arrs = [
        ("End of 2027", rev_summary["arr_2027"]),
        ("End of 2028", rev_summary["arr_2028"]),
        ("End of 2029", rev_summary["arr_2029"]),
    ]
    for i, (year_lbl, arr) in enumerate(arrs):
        r = 7 + i
        ws.cell(row=r, column=1, value=year_lbl).font = _bold()
        ws.cell(row=r, column=2, value=round(arr)).number_format = DOLLAR_FMT
        for mult_i, mult in enumerate([5, 10, 20]):
            val = round(arr * mult)
            ws.cell(row=r, column=3 + mult_i, value=val).number_format = DOLLAR_FMT

    _apply_data_rows(ws, 7, 9, 5, palette)

    # Table 2: Investor ROI for two investment sizes
    invest_amounts = [500000, 1500000]
    table_row = 12

    for inv_amt in invest_amounts:
        inv_pct = inv_amt / (inv_amt + a["pre_money"])  # simplified ownership

        ws.cell(row=table_row, column=1,
                value=f"Investor Returns: ${inv_amt:,.0f} Investment").font = _bold()
        ws.cell(row=table_row + 1, column=1,
                value=f"Ownership: {inv_pct * 100:.1f}% (at ${a['pre_money']:,.0f} pre-money)").font = Font(
            name="Calibri", italic=True, size=10, color="666666"
        )

        inv_headers = ["Exit Scenario", "Company Valuation", "Investor Share",
                       "Return Multiple", "Profit"]
        for c, h in enumerate(inv_headers, 1):
            ws.cell(row=table_row + 3, column=c, value=h)
        _apply_header_row(ws, table_row + 3, 5, palette)

        scenarios = []
        for year_lbl, arr in arrs:
            for mult in [5, 10, 20]:
                company_val = arr * mult
                investor_share = round(company_val * inv_pct)
                return_mult = round(investor_share / inv_amt, 1)
                profit = investor_share - inv_amt
                scenarios.append((
                    f"{year_lbl} @ {mult}x",
                    round(company_val),
                    investor_share,
                    f"{return_mult}x",
                    profit,
                ))

        for i, (scen, val, share, rmult, profit) in enumerate(scenarios):
            r = table_row + 4 + i
            ws.cell(row=r, column=1, value=scen)
            ws.cell(row=r, column=2, value=val).number_format = DOLLAR_FMT
            ws.cell(row=r, column=3, value=share).number_format = DOLLAR_FMT
            ws.cell(row=r, column=4, value=rmult)
            cell_profit = ws.cell(row=r, column=5, value=profit)
            cell_profit.number_format = DOLLAR_FMT
            # Highlight high-return rows
            if profit > inv_amt * 5:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = _kpi_fill(palette)
                    ws.cell(row=r, column=c).font = _bold()

        _apply_data_rows(ws, table_row + 4, table_row + 4 + len(scenarios) - 1, 5, palette)
        table_row += 4 + len(scenarios) + 3

    _auto_col_widths(ws, min_width=16, max_width=35)
    ws.freeze_panes = "A7"


# ===================================================================
# Main workbook builder
# ===================================================================

def build_workbook(assumptions, palette_key, extra_sheets=None):
    """Build a complete financial model workbook.

    extra_sheets: list of callables to add extra sheets, e.g.
        [("scenario", build_scenario_analysis_sheet, {...})]
    """
    a = assumptions
    palette = PALETTES[palette_key]
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Compute all data
    monthly_data = compute_monthly_2026(a)

    end_users_2026 = monthly_data[-1]["total_users"]
    end_restaurants_2026 = monthly_data[-1]["active_restaurants"]

    q2027, end_u_27, end_r_27 = compute_quarterly(a, 2027, end_users_2026, end_restaurants_2026)
    q2028, end_u_28, end_r_28 = compute_quarterly(a, 2028, end_u_27, end_r_27)
    q2029, end_u_29, end_r_29 = compute_quarterly(a, 2029, end_u_28, end_r_28)

    # Build sheets
    rev_summary = build_revenue_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029)
    build_cost_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029)
    build_unit_economics_sheet(wb, a, palette)
    funding_summary = build_funding_sheet(wb, a, palette, monthly_data, q2027, q2028, q2029, rev_summary)
    build_summary_dashboard(wb, a, palette, rev_summary, funding_summary, monthly_data)
    build_assumptions_sheet(wb, a, palette)

    # Move Summary Dashboard to first position
    summary_idx = None
    for i, name in enumerate(wb.sheetnames):
        if name == "Summary Dashboard":
            summary_idx = i
            break
    if summary_idx is not None:
        wb.move_sheet("Summary Dashboard", offset=-summary_idx)

    return wb, monthly_data, q2027, q2028, q2029, rev_summary, funding_summary


def run_full_scenario(assumptions):
    """Run computations and return summary dicts for scenario comparison."""
    monthly_data = compute_monthly_2026(assumptions)
    end_u = monthly_data[-1]["total_users"]
    end_r = monthly_data[-1]["active_restaurants"]
    q2027, eu27, er27 = compute_quarterly(assumptions, 2027, end_u, end_r)
    q2028, eu28, er28 = compute_quarterly(assumptions, 2028, eu27, er27)
    q2029, _, _ = compute_quarterly(assumptions, 2029, eu28, er28)

    rev_2026 = sum(md["total_revenue"] for md in monthly_data)
    rev_2027 = sum(qd["total_revenue"] for qd in q2027)
    rev_2028 = sum(qd["total_revenue"] for qd in q2028)
    rev_2029 = sum(qd["total_revenue"] for qd in q2029)

    arr_2026 = monthly_data[-1]["total_revenue"] * 12
    arr_2027 = q2027[-1]["total_revenue"] * 4
    arr_2028 = q2028[-1]["total_revenue"] * 4
    arr_2029 = q2029[-1]["total_revenue"] * 4

    # Funding/runway
    cash = assumptions["raise_amount"]
    for md in monthly_data:
        cash += md["net"]
    cash_eoy = cash
    avg_burn = sum(md["total_cost"] for md in monthly_data) / 12
    runway = round(cash_eoy / avg_burn, 1) if avg_burn > 0 else 999

    all_q = q2027 + q2028 + q2029
    q_labels = QUARTERS_2027 + QUARTERS_2028 + QUARTERS_2029
    be = "Not in forecast"
    for qi, qd in enumerate(all_q):
        if qd["net"] > 0:
            be = q_labels[qi]
            break

    return {
        "rev": {
            "rev_2026": rev_2026, "rev_2027": rev_2027,
            "rev_2028": rev_2028, "rev_2029": rev_2029,
            "arr_2026": arr_2026, "arr_2027": arr_2027,
            "arr_2028": arr_2028, "arr_2029": arr_2029,
        },
        "fund": {
            "avg_burn": round(avg_burn),
            "cash_eoy_2026": round(cash_eoy),
            "runway_months": runway,
            "break_even": be,
        },
    }


# ===================================================================
# Main
# ===================================================================

def main():
    output_dir = os.path.dirname(os.path.abspath(__file__))

    print("=" * 70)
    print("  FORK Financial Model Generator")
    print("  AI-Powered Dish Intelligence Platform")
    print("=" * 70)

    # ---- Variation 1: Conservative ----
    print("\n[1/3] Building Conservative Model...")
    wb1, md1, q27_1, q28_1, q29_1, rev1, fund1 = build_workbook(CONSERVATIVE, "conservative")
    path1 = os.path.join(output_dir, "fork-financial-model-v1-conservative.xlsx")
    wb1.save(path1)
    size1 = os.path.getsize(path1)
    print(f"      Saved: {path1}")
    print(f"      Size:  {size1:,} bytes ({size1 / 1024:.1f} KB)")
    print(f"      ARR 2026: ${rev1['arr_2026']:,.0f} | 2027: ${rev1['arr_2027']:,.0f} | 2029: ${rev1['arr_2029']:,.0f}")

    # ---- Variation 2: Base Case (with Scenario Analysis) ----
    print("\n[2/3] Building Base Case Model...")
    wb2, md2, q27_2, q28_2, q29_2, rev2, fund2 = build_workbook(BASE_CASE, "base")

    # Run all 3 scenarios for the comparison sheet
    sc_conservative = run_full_scenario(CONSERVATIVE)
    sc_base = run_full_scenario(BASE_CASE)
    sc_aggressive = run_full_scenario(AGGRESSIVE)
    build_scenario_analysis_sheet(wb2, PALETTES["base"], [sc_conservative, sc_base, sc_aggressive])

    path2 = os.path.join(output_dir, "fork-financial-model-v2-base.xlsx")
    wb2.save(path2)
    size2 = os.path.getsize(path2)
    print(f"      Saved: {path2}")
    print(f"      Size:  {size2:,} bytes ({size2 / 1024:.1f} KB)")
    print(f"      ARR 2026: ${rev2['arr_2026']:,.0f} | 2027: ${rev2['arr_2027']:,.0f} | 2029: ${rev2['arr_2029']:,.0f}")

    # ---- Variation 3: Aggressive (with Investor Returns) ----
    print("\n[3/3] Building Aggressive / Upside Model...")
    wb3, md3, q27_3, q28_3, q29_3, rev3, fund3 = build_workbook(AGGRESSIVE, "aggressive")
    build_investor_returns_sheet(wb3, AGGRESSIVE, PALETTES["aggressive"], rev3)

    path3 = os.path.join(output_dir, "fork-financial-model-v3-aggressive.xlsx")
    wb3.save(path3)
    size3 = os.path.getsize(path3)
    print(f"      Saved: {path3}")
    print(f"      Size:  {size3:,} bytes ({size3 / 1024:.1f} KB)")
    print(f"      ARR 2026: ${rev3['arr_2026']:,.0f} | 2027: ${rev3['arr_2027']:,.0f} | 2029: ${rev3['arr_2029']:,.0f}")

    # ---- Summary ----
    print("\n" + "=" * 70)
    print("  ALL FILES GENERATED SUCCESSFULLY")
    print("=" * 70)
    print(f"\n  V1 Conservative : {size1:>8,} bytes  |  {path1}")
    print(f"  V2 Base Case    : {size2:>8,} bytes  |  {path2}")
    print(f"  V3 Aggressive   : {size3:>8,} bytes  |  {path3}")
    print(f"\n  Total           : {size1 + size2 + size3:>8,} bytes")
    print()


if __name__ == "__main__":
    main()
